from openpyxl import load_workbook
from datetime import date

# MAKE SURE CITATIONS.XLSX IS NOT OPEN

# UPDATE THIS AFTER ADDING TO SHEET
num_citations = 12

workbook = load_workbook(filename="c://Users/Jett/Documents/GitHub/dna-ml/Citations.xlsx")
sheet = workbook.active
all_citations = []
labels = [cell.value for cell in sheet[1]]
citation_index = labels.index("Citation") + 1
# print (labels)

for row in sheet.iter_rows(min_row=2):
    citation = {}
    for col in range(citation_index):
        citation[labels[col]] = row[col].value
    all_citations.append(citation)
# print (all_citations)
print (len(all_citations))

def cite(citation): # Authors, [Conference.] Journal Volume(Number), Pages[, Publisher] (Year).
    out = ""
    authors = citation.get("Authors").split("\n")
    year = citation.get("Date").year
    journal = citation.get("Journal")
    title = citation.get("Title")
    volume = citation.get("Vol.")
    number = citation.get("No.", None)
    pages = citation.get("pp.", None)
    conference = citation.get("Conference", None)
    publisher = citation.get("Publisher", None)
    doi = citation.get("DOI")
    for a in authors:
        names = a.split(" ")
        lastname = names[-1]
        for name in names[:-1]:
            out += name[0] + ". "
        out += lastname 
        if len(authors) > 4: 
            out += " et al., "
            break
        out += ", "
    if conference is not None:
        out += str(conference) + ". "
    out += journal + " "
    out += str(volume)
    if number is not None:
        out += "(" + str(number) + "), "
    else:
        out += ", "
    if pages is not None:
        out += str(pages)
    if publisher is not None:
        out += ", " + str(publisher)
    out += " (" + str(year) + ")."
    return out

for row in range(2, 2+num_citations):
    sheet.cell(row=row, column=citation_index).value = cite(all_citations[row-2])
    print (cite(all_citations[row-2]))

workbook.save(filename="c://Users/Jett/Documents/GitHub/dna-ml/Citations.xlsx")
    