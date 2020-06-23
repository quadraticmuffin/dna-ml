import requests
r = requests.get('https://www.fpbase.org/api/proteins/spectra/?format=json')

import json
# with open('spectra.json', 'w', encoding='utf-8') as f:
#     json.dump(r.json(), f, ensure_ascii=False, indent=4)
from openpyxl import Workbook
# TODO ask API authors for their data on non-protein dyes that are not in API

# Take into account:
# - spectral overlap
# - distance between emission spectra
# - quantum yield?
# - extinction coeff?

# Here I'm only concerned with Forster radius. Ideally we also minimize the overlap
# between donor emission and acceptor emission, but that's not my goal with this code.

# Naive fret function, numerically integrates spectra then multiplies by qy and ec
def fret (em, ex):
    emArray = em["data"] # array of [wavelength, intensity] arrays
    exArray = ex["data"]

    minLam = int(max(emArray[0][0], exArray[0][0]))
    maxLam = int(min(emArray[-1][0], exArray[-1][0]))
    emIdx = 0
    while emArray[emIdx][0] < minLam:
        emIdx += 1
    exIdx = 0
    while exArray[exIdx][0] < minLam:
        exIdx += 1

    J = 0    
    for lam in range(minLam, maxLam+1):
        J += emArray[emIdx][1] * exArray[exIdx][1]
        emIdx += 1
        exIdx += 1
    J *= em["qy"] * ex["ec"]
    return J

with open('FPBase Data.json', 'r', encoding='utf-8') as f:
    fpbase_data = json.load(f)

def test1(): 
    fluor1 = data[0]["spectra"]
    fluor2 = data[1]["spectra"]
    # ex is 0, em is 1
    print ("Ex:" + data[0]["name"])
    print ("Em:" + data[1]["name"])
    print ("Relative efficiency: " + str(fret(fluor2[1], fluor1[0])))
    print ("Ex:" + data[1]["name"])
    print ("Em:" + data[0]["name"])
    print ("Relative efficiency: " + str(fret(fluor1[1], fluor2[0])))
    # expect one of these to be zero
# Realized that some fluors in data have more than two spectra,
# accounting for multiple states

# Checked that I can separate the spectra into excitation and emission 
# since no spectra states have both or neither of "ex", "em"
def verify_separation(data):
    for fluor in data:
        for spectrum in fluor["spectra"]:
            state = spectrum["state"]
            if state.find("em") >= 0 and state.find("ex") >= 0:
                print (fluor["name"])
                print (state)
            if state.find("em") < 0 and state.find("ex") < 0:
                print (fluor["name"])
                print (state)

def clean(spectrum):
    arr = [0] * 1000
    for item in spectrum:
        arr[int(item[0])] = item[1]
    return arr

# generate separate JSON lists for emission and excitation data from FPBase
# Mostly (if not all) fluorescent proteins, few to no organic dyes etc.
def separate(data):
    em = []
    ex = []
    for fluor in data:
        for spectrum in fluor["spectra"]:
            state = spectrum["state"]
            newItem = {}
            newItem["name"] = fluor["name"]
            newItem["state"] = state
            newItem["max"] = spectrum["max"]

            if state.find("em") >= 0:
                newItem["qy"] = spectrum["qy"]
                newItem["data"] = spectrum["data"]
                em.append(newItem)

            else: # if not emission, must be excitation
                # newItem["ec"] = spectrum["ec"]
                newItem["data"] = spectrum["data"]
                ex.append(newItem)

    def getMax(spectrum):
        if spectrum["max"] is None:
            return 65535
        return spectrum["max"]

    em.sort(key=getMax)
    with open('FPBase Emission.json', 'w', encoding='utf-8') as f:
        json.dump(em, f, ensure_ascii=False, indent=4)
    print("FPBase Emission done")
    ex.sort(key=getMax)
    with open('FPBase Excitation.json', 'w', encoding='utf-8') as f:
        json.dump(ex, f, ensure_ascii=False, indent=4)
    print ("FPBase Excitation done") # TODO are em and ex flipped? 


# From Arizona 
# Cleaned files: turned into JSON format, removed some NULL entries

def arizona_compile_and_separate():
    arizona_data = [{} for i in range(2000)]

    with open('Arizona Items.json', 'r', encoding='utf-8') as f:
        arizona_items = json.load(f)
    dye_ids = [False] * 2000
    for item in arizona_items:
        if item[2] == "dye":
            dye_ids[item[0]] = True
            arizona_data[item[0]]["name"] = item[1]
            arizona_data[item[0]]["state"] = item[3]
    del(arizona_items)

    with open('Arizona Info.json', 'r', encoding='utf-8') as f:
        arizona_info = json.load(f)
    for item in arizona_info:
        if dye_ids[item[0]]:
            arizona_data[item[0]][item[1]] = item[2]
    del(arizona_info)

    with open('Arizona Spectra.json', 'r', encoding='utf-8') as f:
        arizona_spectra = json.load(f)

    for fluor in arizona_data:
            fluor["data"] = []
        
    for item in arizona_spectra:
        if dye_ids[item[0]]:
            if item[1] < 1000:
                # TODO fix this... do i really need to worry about any wavelengths over 1000 nm?
                # and I don't know what unit is being used when item[2] goes over 1.
                # It's definitely not relative intensity anymore.
                arizona_data[item[0]]["data"].append([item[1], item[2]])
        # arizona_spectra contains non-integer wavelengths; I ignore and round.
        # jk this might be a bad idea, for the spectra that also skip wavelengths. 
        # I think I have to just make a smarter Forster radius function
        # which accounts for this with thicker/thinner Riemann sum.
    del(arizona_spectra)
    for i in range(len(arizona_data)-1, -1, -1):
        if arizona_data[i]["data"] == []:
            del(arizona_data[i])
    with open('Arizona Data.json', 'w', encoding='utf-8') as f:
        json.dump(arizona_data, f, ensure_ascii=False, indent=4)
    
    em = []
    ex = []
    for fluor in arizona_data:
        state = fluor["state"]
        newItem = {}
        newItem["name"] = fluor["name"]
        newItem["state"] = state
        # newItem["max"] = fluor["max"] #TODO This goes by different names in Arizona
        no_qy_dyes = []
        if state == "EM":
            if "Quantum Yield" not in fluor or fluor["Quantum Yield"] == "XXX":
                # print (fluor["name"])
                newItem["qy"] = None # TODO how do i track down all these qy's?
                no_qy_dyes.append(fluor["name"])
            else:
                newItem["qy"] = fluor["Quantum Yield"]
            newItem["data"] = fluor["data"]
            em.append(newItem)

        else: # if not emission, must be excitation/absorbance
            newItem["data"] = fluor["data"]
            ex.append(newItem)
    with open('Dyes with no QY.txt', 'w', encoding='utf-8') as f:
        json.dump(no_qy_dyes, f, ensure_ascii=False, indent=4)
    # def getMax(fluor):
    #     if fluor["max"] is None:
    #         return 12345
    #     return fluor["max"]

    # em.sort(key=getMax)
    with open('Arizona Emission.json', 'w', encoding='utf-8') as f:
        json.dump(em, f, ensure_ascii=False, indent=4)
    print("Arizona Emission done")
    # ex.sort(key=getMax)
    with open('Arizona Excitation.json', 'w', encoding='utf-8') as f:
        json.dump(ex, f, ensure_ascii=False, indent=4)

    print ("Arizona done")

# TODO replace interpolate, integrate with scipy

def interpolate(a, b):
    a.sort(key=lambda x: x[0])
    b.sort(key=lambda x: x[0])
    floor = max(a[0][0], b[0][0])
    ceil = min(a[-1][0], b[-1][0])
    ai = 0
    bi = 0

    if ceil <= floor: return [[0,0]],[[0,0]]

    while a[ai][0] < floor:
        ai += 1
    while b[bi][0] < floor:
        bi += 1
    
    while a[ai][0] < ceil or b[bi][0] < ceil:
        if a[ai][0] > b[bi][0]: 
            a.append([b[bi][0], a[ai-1][1] + (a[ai][1] - a[ai-1][1])
            *(b[bi][0]-a[ai-1][0])
            /(a[ai][0]-a[ai-1][0])])
            bi += 1
        elif b[bi][0] > a[ai][0]:
            b.append([a[ai][0], b[bi-1][1] + (b[bi][1] - b[bi-1][1])
            *(a[ai][0]-b[bi-1][0])
            /(b[bi][0]-b[bi-1][0])])
            ai += 1
        else: 
            ai += 1
            bi += 1
            
    
    if a[ai][0] > b[bi][0]:
        a.append([b[bi][0], a[ai-1][1] + (a[ai][1] - a[ai-1][1])
        *(b[bi][0]-a[ai-1][0])
        /(a[ai][0]-a[ai-1][0])])
    elif b[bi][0] > a[ai][0]:
        b.append([a[ai][0], b[bi-1][1] + (b[bi][1] - b[bi-1][1])
        *(a[ai][0]-b[bi-1][0])
        /(b[bi][0]-b[bi-1][0])])

    a.sort(key=lambda x: x[0])
    b.sort(key=lambda x: x[0])
    while a[0][0] < floor:
        del(a[0])
    while a[-1][0] > ceil:
        del(a[-1])
    while b[0][0] < floor:
        del(b[0])
    while b[-1][0] > ceil:
        del(b[-1])
    return a,b

def total(em):
    out = 0
    for i in range(len(em) - 1):
        step_size = em[i+1][0] - em[i][0]
        out += step_size * (em[i][1] + em[i+1][1])/2
    return out

import scipy.integrate
import numpy as np
import pandas as pd
PI = np.pi
AVOGADRO = 6.022e23 # molecules / mole
LIGHT_SPEED = 3e8 # m/s
# Forster radius
#TODO brUH WHAT UNITS DOES THIS OUTPUT
def calc_forster_radius(donor_emission = None, acceptor_absorbance = None, kappa2 = 2./3, quant_yield = 1.0, refractive_index = 1.4):
    # donor_emission: 1D numpy array, units of inverse distance
    # acceptor_absorbance: 1D numpy array, in terms of wavelength, units of inverse concentration * inverse distance
    # kappa2 (default 2/3): float, unitless
    # quant_yield (default 1.0): float, unitless
    # refractive_index: float, unitless

    if quant_yield is None: #TODO fix this abomination? why are some of the quantum yields 0
        return -1
    coeff = 0
    try:
        coeff = 9. * np.log(10) * kappa2 * float(quant_yield) / (128. * PI**5 * AVOGADRO * refractive_index**4)
    except TypeError:
        print (type(quant_yield), quant_yield)
    if total(donor_emission) == 0:
        return 0
    em, ab = interpolate(donor_emission, acceptor_absorbance)
    
    # integrate
    integral = 0
    
    for i in range(len(em) - 1):
        step_size = em[i+1][0] - em[i][0]
        lam = em[i][0]
        
        integral += step_size * lam**4 * em[i][1]/total(em) * ab[i][1]
    # TODO assert integral >= 0, "negative integral"
    if integral < 0:
        return -1
    return (coeff * integral)**(1./6)

# a = [[1, 1],[2,2],[3,3],[5,5],[6,6]]
# b = [[0,2],[0.5,2],[1.25,5],[1.75,6],[2,5],[2.25,4],[2.5,2]]
# print (interpolate(a,b)[0])
# print (interpolate(a,b)[1])

def combine():
    with open('FPBase Emission.json', 'r', encoding='utf-8') as f:
        emission = json.load(f)
    with open('FPBase Excitation.json', 'r', encoding='utf-8') as f:
        excitation = json.load(f)
    with open('Arizona Emission.json', 'r', encoding='utf-8') as f:
        emission.extend(json.load(f))
    with open('Arizona Excitation.json', 'r', encoding='utf-8') as f:
        excitation.extend(json.load(f))
    with open('Combined Emission.json', 'w', encoding='utf-8') as f:
        json.dump(emission, f, ensure_ascii=False, indent=4)
    with open('Combined Excitation.json', 'w', encoding='utf-8') as f:
        json.dump(excitation, f, ensure_ascii=False, indent=4)
    return

def main():
    with open('Combined Emission.json', 'r', encoding='utf-8') as f:
        emission = json.load(f)
    with open('Combined Excitation.json', 'r', encoding='utf-8') as f:
        excitation = json.load(f)
    # wb = Workbook()
    dest_filename = 'results.xlsx'
    # dest = wb.active
    # dest.title = "Forster Radii"

    # results = [0] * 1000000
    radii = np.empty([1100, 1100], dtype="S20")
    # i=0
    radii[0,0] = "donors\\acceptors"
    for i in range(len(emission)):
        donor = emission[i]
        radii[i+1,0] = donor["name"]
        # _ = dest.cell(row=i+2, column=1, value = donor["name"])
        for j in range(len(excitation)):
            
            acceptor = excitation[i]
            #  _ = dest.cell(row=i+2, column=j+2, 
            # value=calc_forster_radius(donor["data"], acceptor["data"], quant_yield=donor["qy"]))
            radii[i+1,j+1] = calc_forster_radius(donor["data"], acceptor["data"], quant_yield=donor["qy"])
            # try:
            #     result = ("donor: " + donor["name"],"acceptor: " + acceptor["name"], 
            #     # "Avg wavelength: " + str(donor["max"]/2 + acceptor["max"]/2), 
            #     calc_forster_radius(donor["data"], acceptor["data"], quant_yield=donor["qy"]))
                
            # except AssertionError:
            #     print (donor["name"], acceptor["name"])
            # results[i] = result
            # i += 1
        print ("inserted values " + str(i) + "/" + str(len(emission)))
    for j in range(len(excitation)):
        acceptor = excitation[j]
        radii[0,j+1] = acceptor["name"]
        # _ = dest.cell(row=1, column=j+2, value = acceptor["name"])
    def temp(x):
        if isinstance(x, int):
            return -1000000
        return -x[2]
    # results.sort(key=temp)
    # results = results[results.count(0):]
    # with open('Results.json', 'w', encoding='utf-8') as f:
    #     json.dump(results, f, ensure_ascii=False, indent=4)
    # wb.save(filename = dest_filename)
    print("converting to DataFrame")
    df = pd.DataFrame(data=radii[1:,1:],    # values
        index=radii[1:,0],    # 1st column as index
        columns=radii[0,1:])  # 1st row as the column names

    print ("Saving to Excel")
    df.to_excel(dest_filename, index=False)

    print("done ayyyyy")

# separate(fpbase_data)
# del(fpbase_data)
# arizona_compile_and_separate()
# combine()
# print (calc_forster_radius())
# main()

def test():
    with open('Arizona Emission.json', 'r', encoding='utf-8') as f:
        arizona_emission = json.load(f)
    with open('Arizona Excitation.json', 'r', encoding='utf-8') as f:
        arizona_excitation = json.load(f)
    return arizona_emission, arizona_excitation

import matplotlib.pyplot as plt
def plot(spectra):
    for spectrum in spectra:
        lam,y = zip(*spectrum)
        plt.plot(lam, y)
    plt.show()
